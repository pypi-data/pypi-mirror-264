#!/usr/bin/env python3
""" certwatch - watch X509 certificates expiration dates
License: 3-clause BSD (see https://opensource.org/licenses/BSD-3-Clause)
Author: Hubert Tournier
"""

import datetime
import logging
import os
import platform
import re
import signal
import socket
import ssl
import subprocess
import time

import cryptography.x509
import cryptography.hazmat.backends
import cryptography.hazmat.primitives
import openpyxl
import openpyxl.styles
import prettytable
import tqdm

DEFAULT_PORT = 443 # HTTPS port

####################################################################################################
def read_input_file(filename):
    """ Get the contents of the file as a list of lines
    A FileNotFoundError exception could be raised and should be caught in the calling function
    """
    with open(filename, encoding="utf-8") as file:
        lines = file.read().splitlines()

    return lines


####################################################################################################
def _process_input_line(line):
    """ Process a line of the input file and returns hostname/hostport information if present
    The input line can be a blank line, a comment line (starting with #) or a data line of the form:
        name [port] [# comment]
    """
    line = re.sub(r"#.*", "", line)
    line = line.strip()
    words = line.split()
    if len(words) == 0:
        raise Warning
    if len(words) == 1:
        hostname = words[0]
        hostport = DEFAULT_PORT
    elif len(words) == 2:
        hostname = words[0]
        try:
            hostport = int(words[1])
        except ValueError as error:
            raise TypeError from error
        if hostport < 0 or hostport > 65535:
            raise ValueError
    else:
        raise SyntaxError

    return hostname, hostport


####################################################################################################
def get_input_data(lines):
    """ Returns a list of hostname / hostport information from the input lines """
    input_data = []
    for line in lines:
        try:
            hostname, hostport = _process_input_line(line)
        except Warning:
            # empty line
            continue
        except TypeError:
            logging.error("invalid port number in line '%s'", line)
            continue
        except ValueError:
            logging.error("invalid port value in line '%s'", line)
            continue
        except SyntaxError:
            logging.error("invalid data line '%s'", line)
            continue

        input_data.append([hostname, hostport])

    return input_data


####################################################################################################
def _add_items_to_dict(target_dict, extension):
    """ Add an extension to a dictionary mapping a certificate
    A true parser would be better, but this one will do the job for the time being.
    I'm working on a base64 -> ASN.1 -> Python dict decoder to replace the "cryptography" module
    whose API is not fully satisfying.
    Everything that is not used by this software is commented
    """
    group = re.match(r"^.*, name=([A-Za-z]+)\).*, value=<[A-Za-z]+\((.+)\)>\)>$", str(extension))
    if group is None:
        return

    if group[2].startswith("<GeneralNames(["):
        value = re.sub(r"<GeneralNames\(\[", "", group[2])
        value = re.sub(r"\]\)>", "", value)
        value = re.sub(r"<DNSName\(value='", "", value)
        value = re.sub(r"'\)>", "", value)
        target_dict[group[1]] = []
        for part in value.split(", "):
            target_dict[group[1]].append(part)
    #pylint: disable=W1401
    """
    elif group[2].startswith("["):
        if group[1] == "authorityInfoAccess":
            name = ""
            value = re.sub(r"^\[", "", group[2])
            value = re.sub(r"\]$", "", value)
            target_dict[group[1]] = {}
            for part in value.split(", "):
                if part.startswith("name="):
                    part = re.sub(r"^name=", "", part)
                    name = re.sub(r"\).*$", "", part)
                elif part.startswith("access_location"):
                    part = re.sub(r"^access_location=<UniformResourceIdentifier\(value='", "", part)
                    part = re.sub(r"'.*$", "", part)
                    target_dict[group[1]][name] = part
        elif group[1] == "certificatePolicies":
            # Unmanaged!
            target_dict[group[1]] = group[2]
        elif group[1] == "extendedKeyUsage":
            value = re.sub(r"^\[", "", group[2])
            value = re.sub(r"\]$", "", value)
            target_dict[group[1]] = []
            for part in value.split(", "):
                if part.startswith("name="):
                    part = re.sub(r"^name=", "", part)
                    part = re.sub(r"\).*$", "", part)
                    target_dict[group[1]].append(part)
        elif group[1] == "signedCertificateTimestampList":
            target_dict[group[1]] = "A list of (uncollected) signed certificate timestamps"
        else:
            # Unmanaged!
            target_dict[group[1]] = group[2]
    else:
        target_dict[group[1]] = {}
        for part in group[2].split(", "):
            key = part.split("=")[0]
            value = part.split("=")[1]
            if value in ["True", "False"]:
                target_dict[group[1]][key] = value == "True"
            elif value == "None":
                target_dict[group[1]][key] = None
            else:
                target_dict[group[1]][key] = value
    """
    #pylint: enable=W1401


####################################################################################################
def _decode_pem_cert(pem_cert):
    """ Decode a PEM certificate into a Python dict
    We returned far more information than needed, which is not that useful finally, as the
    "cryptography" module doesn't seem to offer a consistent API to fully decode a certificate,
    so all the unused parts are now commented...
    """
    certificate = {}

    cert = cryptography.x509.load_pem_x509_certificate(str.encode(pem_cert), cryptography.hazmat.backends.default_backend())

    #version = {}
    #version[cert.version.name] = cert.version.value
    #certificate["version"] = version

    #certificate["fingerprint"] = "0x" + cert.fingerprint(cryptography.hazmat.primitives.hashes.SHA256()).hex()
    #certificate["serial number"] = cert.serial_number

    #public_key = cert.public_key()

    #certificate["not valid before"] = cert.not_valid_before_utc
    certificate["not valid after"] = cert.not_valid_after_utc

    issuer = {}
    for attribute in cert.issuer:
        issuer[attribute.oid._name] = attribute.value # who needs privacy anyway?
    certificate["issuer"] = issuer

    subject = {}
    for attribute in cert.subject:
        subject[attribute.oid._name] = attribute.value # who needs privacy anyway?
    certificate["subject"] = subject

    #print(f"signature_hash_algorithm : {cert.signature_hash_algorithm}")
    #certificate["signature algorithm"] = cert.signature_algorithm_oid._name
    #print(f"signature_algorithm_parameters : {cert.signature_algorithm_parameters}")

    certificate["extensions"] = {}
    for extension in cert.extensions:
        _add_items_to_dict(certificate["extensions"], extension)

    #certificate["signature"] = "0x" + cert.signature.hex()

    return certificate


####################################################################################################
def _get_cert(hostname, hostport, save_cert_dir):
    """ Get a certificate for hostname on the given port
    Works only on TLS servers, as Server Name Indication (SNI) functionality is required.
    """
    try:
        connection = ssl.create_connection((hostname, hostport))
    except (socket.gaierror, OSError) as error:
        error_message = re.sub(r".*] ", "", str(error))
        raise NameError(error_message) from error
    # There can also be a TimeoutError which needs to be caught in the calling function

    # Despite the name, the following line auto-negotiate the highest protocol version
    # that both the client and server support. PROTOCOL_TLS is noted as deprecated, but
    # its replacement PROTOCOL_TLS_CLIENT implies additional controls that we don't want...
    context = ssl.SSLContext(ssl.PROTOCOL_TLS)
    try:
        sock = context.wrap_socket(connection, server_hostname=hostname)
    except ssl.SSLError as error:
        error_message = re.sub(r".*] ", "", str(error))
        error_message = re.sub(r" \([^)]*\)$", "", error_message)
        raise ConnectionError(error_message) from error

    der_cert = sock.getpeercert(True)
    pem_cert = ssl.DER_cert_to_PEM_cert(der_cert)
    sock.close()

    if save_cert_dir:
        os.makedirs(save_cert_dir, exist_ok=True)
        filename = save_cert_dir + os.sep + hostname + "_" + str(hostport) + ".pem"
        with open(filename, "w", encoding="utf-8") as file:
            file.write(pem_cert)

    return _decode_pem_cert(pem_cert)


####################################################################################################
def _timeout_signal_handler(signum, frame):
    """ Connection timeout signal handler """
    raise TimeoutError


####################################################################################################
def _resolve_hostname(hostname):
    """ Resolve the hostname into an IP address or raise a NameError exception """
    try:
        ip_address = socket.gethostbyname(hostname)
    except socket.gaierror as error:
        error_message = re.sub(r".*] ", "", str(error))
        raise NameError(error_message) from error

    return ip_address


####################################################################################################
def _is_pingable(hostname):
    """ Use the system ping command to (try to) check if an hostname is up """
    if platform.system().lower() == "windows":
        option = "-n"
    else:
        option = "-c"

    command = [ "ping", option, "1", hostname]
    return subprocess.run(
        args = command,
        stdout = subprocess.DEVNULL,
        stderr = subprocess.DEVNULL,
        check = False,
        shell = True,
    ).returncode == 0


####################################################################################################
def _is_listening(hostname, hostport, timeout=10):
    """ Attempts a connection on the given port and reports if it's open or not """
    sock = socket.socket()
    if platform.system().lower() != "windows":
        signal.signal(signal.SIGALRM, _timeout_signal_handler)
        signal.alarm(timeout)
    try:
        sock.connect((hostname, hostport))
        sock.close()
        listening = True
    except socket.error:
        listening = False
    finally:
        if platform.system().lower() != "windows":
            signal.alarm(0)

    return listening


####################################################################################################
def _analyze_error(hostname, hostport, ip_address, message, exception):
    """ Analyze the possible reasons for not getting a certificate """
    if not ip_address:
        ping = "N/A"
        listening = "N/A"
        listen80 = "N/A"
    else:
        if _is_pingable(hostname):
            ping = "YES"
        else:
            ping = "NO"

        if _is_listening(hostname, hostport):
            listening = "YES"
        else:
            listening = "NO"

        # The following is probably not useful anymore
        if not message:
            if exception == "NameError":
                message = "Connection failed"
            elif exception == "TimeoutError":
                message = "Connection timeout"
            elif exception == "ConnectionError":
                message = "SSL/TLS handshake failed"

        if _is_listening(hostname, 80):
            listen80 = "YES"
        else:
            listen80 = "NO"

    return {"IP": ip_address, "Ping": ping, "Listening": listening, "Message": message, "Listen:80": listen80}


####################################################################################################
def get_certs(input_data, progress_bar=True, delay=1, timeout=10, save_cert_dir=""):
    """ Return dictionaries of certificates fetched and errors encountered """
    certs = {}
    errors = {}
    if platform.system().lower() != "windows":
        signal.signal(signal.SIGALRM, _timeout_signal_handler)
    if progress_bar:
        for i in tqdm.tqdm(range(len(input_data))):
            hostname = input_data[i][0]
            hostport = input_data[i][1]

            error_message = ""
            try:
                ip_address = _resolve_hostname(hostname)
            except NameError as error:
                ip_address = ""
                error_message = str(error)

            if ip_address:
                if platform.system().lower() != "windows":
                    signal.alarm(timeout)
                try:
                    cert = _get_cert(hostname, hostport, save_cert_dir)
                except (NameError, TimeoutError, ConnectionError) as error:
                    if platform.system().lower() != "windows":
                        signal.alarm(0)
                    errors[hostname + ":" + str(hostport)] = _analyze_error(hostname, hostport, ip_address, str(error), type(error).__name__)
                    continue
                if platform.system().lower() != "windows":
                    signal.alarm(0)

                cert["IP"] = ip_address
                certs[hostname + ":" + str(hostport)] = cert

                # Avoid doing a denial of service on target machines by spitting requests to fast
                time.sleep(delay)
            else:
                errors[hostname + ":" + str(hostport)] = _analyze_error(hostname, hostport, "", error_message, "")
    else:
        for data in input_data:
            hostname = data[0]
            hostport = data[1]

            error_message = ""
            try:
                ip_address = _resolve_hostname(hostname)
            except NameError as error:
                ip_address = ""
                error_message = str(error)

            if ip_address:
                if platform.system().lower() != "windows":
                    signal.alarm(timeout)
                try:
                    cert = _get_cert(hostname, hostport, save_cert_dir)
                except (NameError, TimeoutError, ConnectionError) as error:
                    if platform.system().lower() != "windows":
                        signal.alarm(0)
                    errors[hostname + ":" + str(hostport)] = _analyze_error(hostname, hostport, ip_address, str(error), type(error).__name__)
                    continue
                if platform.system().lower() != "windows":
                    signal.alarm(0)

                cert["IP"] = ip_address
                certs[hostname + ":" + str(hostport)] = cert

                # Avoid doing a denial of service on target machines by spitting requests to fast
                time.sleep(delay)
            else:
                errors[hostname + ":" + str(hostport)] = _analyze_error(hostname, hostport, "", error_message, "")

    return certs, errors


####################################################################################################
def print_table(certs, errors, show_alt_names=False, show_ip=False, expiration=None):
    """ Print a table of certificates ordered by expiration date and a list of errors encountered """
    nb_certs = 0
    t = prettytable.PrettyTable()
    if show_alt_names:
        if show_ip:
            t.field_names = ["hostname:port", "IP address", "common name", "alt names", "issuer org name", "not valid after"]
        else:
            t.field_names = ["hostname:port", "common name", "alt names", "issuer org name", "not valid after"]
    else:
        if show_ip:
            t.field_names = ["hostname:port", "IP address", "common name", "issuer org name", "not valid after"]
        else:
            t.field_names = ["hostname:port", "common name", "issuer org name", "not valid after"]
    t.align = "l"
    t.sortby = "not valid after"
    t.set_style(prettytable.SINGLE_BORDER)

    for key, value in certs.items():
        try:
            common_name = value['subject']['commonName']
        except KeyError:
            common_name = ""
        try:
            alt_names = value['extensions']['subjectAltName']
        except KeyError:
            alt_names = []
        try:
            issuer = value['issuer']['organizationName']
        except KeyError:
            issuer = ""
        try:
            not_valid_after = value['not valid after']
        except KeyError:
            not_valid_after = datetime.datetime(1, 1, 1, 0, 0, 0, 0, tzinfo=datetime.timezone.utc)

        if expiration is not None:
            if not_valid_after > datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(days=expiration):
                # We have all the time in the world...
                continue
            nb_certs += 1

        all_alt_names = ""
        if len(alt_names):
            for alt_name in alt_names:
                all_alt_names += alt_name + "\n"
            #all_alt_names = all_alt_names[:-1]

        if show_alt_names:
            if show_ip:
                t.add_row([key, value["IP"], common_name, all_alt_names, issuer, not_valid_after])
            else:
                t.add_row([key, common_name, all_alt_names, issuer, not_valid_after])
        elif show_ip:
            t.add_row([key, value["IP"], common_name, issuer, not_valid_after])
        else:
            t.add_row([key, common_name, issuer, not_valid_after])

    e = prettytable.PrettyTable()
    e.field_names = ["hostname", "IP address", "is pingable?", "port", "is listening?", "error type", "is listening on http?"]
    e.align = "l"
    e.sortby = "hostname"
    e.set_style(prettytable.SINGLE_BORDER)

    for key, value in errors.items():
        hostname = key.split(":")[0]
        hostport = int(key.split(":")[1])
        e.add_row([hostname, value["IP"], value["Ping"], hostport, value["Listening"], value["Message"], value["Listen:80"]])

    if expiration is None:
        print(f"\nCertificates expiration dates: (showing all {len(certs)})")
    else:
        print(f"\nCertificates expiration dates in less than {expiration} days: (showing {nb_certs}/{len(certs)})")
    print(t)

    if len(errors):
        print(f"\nServer errors: (showing all {len(errors)})")
        print(e)


####################################################################################################
def generate_excel(excel_filename, certs, errors, expiration=None):
    """ Generate an Excel workbook with tabs for certificates and errors encountered """
    excel_workbook = openpyxl.Workbook()

    excel_sheet = excel_workbook.active
    excel_sheet.page_setup.orientation = excel_sheet.ORIENTATION_LANDSCAPE
    excel_sheet.page_setup.fitToWidth = 1
    excel_line = 1
    max_width = {}

    excel_sheet.title = "Expiration dates"
    excel_sheet["A1"] = "hostname:port"
    excel_sheet["A1"].font = openpyxl.styles.Font(bold=True)
    max_width["A"] = len("hostname:port")
    excel_sheet["B1"] = "IP address"
    excel_sheet["B1"].font = openpyxl.styles.Font(bold=True)
    max_width["B"] = len("IP address")
    excel_sheet["C1"] = "common name"
    excel_sheet["C1"].font = openpyxl.styles.Font(bold=True)
    max_width["C"] = len("common name")
    excel_sheet["D1"] = "alt names"
    excel_sheet["D1"].font = openpyxl.styles.Font(bold=True)
    max_width["D"] = len("alt names")
    excel_sheet["E1"] = "issuer org name"
    excel_sheet["E1"].font = openpyxl.styles.Font(bold=True)
    max_width["E"] = len("issuer org name")
    excel_sheet["F1"] = "not valid after"
    excel_sheet["F1"].font = openpyxl.styles.Font(bold=True)
    max_width["F"] = len("not valid after")
    excel_sheet.freeze_panes = "A2"

    for key, value in certs.items():
        try:
            common_name = value['subject']['commonName']
        except KeyError:
            common_name = ""
        try:
            alt_names = value['extensions']['subjectAltName']
        except KeyError:
            alt_names = []
        try:
            issuer = value['issuer']['organizationName']
        except KeyError:
            issuer = ""
        try:
            not_valid_after = value['not valid after']
        except KeyError:
            not_valid_after = datetime.datetime(1, 1, 1, 0, 0, 0, 0, tzinfo=datetime.timezone.utc)

        all_alt_names = ""
        if len(alt_names):
            for alt_name in alt_names:
                all_alt_names += alt_name + "\n"
            all_alt_names = all_alt_names[:-1]

        color = "000000" # black
        if not_valid_after < datetime.datetime.now(datetime.timezone.utc): # expired
            color = "800000" # crimson
        elif expiration is not None:
            if not_valid_after < datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(days=expiration):
                color = "FF0000" # red

        excel_line += 1
        excel_sheet["A" + str(excel_line)] = key
        excel_sheet["A" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top")
        excel_sheet["A" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(key) > max_width["A"]:
            max_width["A"] = len(key)
        excel_sheet["B" + str(excel_line)] = value["IP"]
        excel_sheet["B" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top")
        excel_sheet["B" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["IP"]) > max_width["B"]:
            max_width["B"] = len(value["IP"])
        excel_sheet["C" + str(excel_line)] = common_name
        excel_sheet["C" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top")
        excel_sheet["C" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(common_name) > max_width["C"]:
            max_width["C"] = len(common_name)
        excel_sheet["D" + str(excel_line)] = all_alt_names
        excel_sheet["D" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top", wrapText=True)
        excel_sheet["D" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        for line in all_alt_names.split("\n"):
            if len(line) > max_width["D"]:
                max_width["D"] = len(line)
        excel_sheet["E" + str(excel_line)] = issuer
        excel_sheet["E" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top")
        excel_sheet["E" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(issuer) > max_width["E"]:
            max_width["E"] = len(issuer)
        excel_sheet["F" + str(excel_line)] = str(not_valid_after)
        excel_sheet["F" + str(excel_line)].alignment = openpyxl.styles.Alignment(vertical="top")
        excel_sheet["F" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(str(not_valid_after)) > max_width["F"]:
            max_width["F"] = len(str(not_valid_after))
        excel_sheet.row_dimensions[excel_line].height = 15 * (1 + all_alt_names.count("\n"))

    excel_sheet.column_dimensions["A"].width = max_width["A"] + 1
    excel_sheet.column_dimensions["B"].width = max_width["B"] + 1
    excel_sheet.column_dimensions["C"].width = max_width["C"] + 1
    excel_sheet.column_dimensions["D"].width = max_width["D"] + 1
    excel_sheet.column_dimensions["E"].width = max_width["E"] + 1
    excel_sheet.column_dimensions["F"].width = max_width["F"] + 1


    excel_workbook.create_sheet(index=1, title="Errors")
    excel_sheet = excel_workbook.get_sheet_by_name("Errors")
    excel_sheet.page_setup.orientation = excel_sheet.ORIENTATION_LANDSCAPE
    excel_sheet.page_setup.fitToWidth = 1
    excel_line = 1
    max_width = {}

    excel_sheet["A1"] = "hostname"
    excel_sheet["A1"].font = openpyxl.styles.Font(bold=True)
    max_width["A"] = len("hostname")
    excel_sheet["B1"] = "IP address"
    excel_sheet["B1"].font = openpyxl.styles.Font(bold=True)
    max_width["B"] = len("IP address")
    excel_sheet["C1"] = "is pingable?"
    excel_sheet["C1"].font = openpyxl.styles.Font(bold=True)
    max_width["C"] = len("is pingable?")
    excel_sheet["D1"] = "port"
    excel_sheet["D1"].font = openpyxl.styles.Font(bold=True)
    max_width["D"] = len("port")
    excel_sheet["E1"] = "is listening?"
    excel_sheet["E1"].font = openpyxl.styles.Font(bold=True)
    max_width["E"] = len("is listening?")
    excel_sheet["F1"] = "error type"
    excel_sheet["F1"].font = openpyxl.styles.Font(bold=True)
    max_width["F"] = len("error type")
    excel_sheet["G1"] = "is listening on http?"
    excel_sheet["G1"].font = openpyxl.styles.Font(bold=True)
    max_width["G"] = len("is listening on http?")
    excel_sheet.freeze_panes = "A2"

    for key, value in errors.items():
        hostname = key.split(":")[0]
        hostport = int(key.split(":")[1])

        color = "000000" # black
        if not value["IP"]:
            color = "808080" # gray

        excel_line += 1
        excel_sheet["A" + str(excel_line)] = hostname
        excel_sheet["A" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(hostname) > max_width["A"]:
            max_width["A"] = len(hostname)
        excel_sheet["B" + str(excel_line)] = value["IP"]
        excel_sheet["B" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["IP"]) > max_width["B"]:
            max_width["B"] = len(value["IP"])
        excel_sheet["C" + str(excel_line)] = value["Ping"]
        excel_sheet["C" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["Ping"]) > max_width["C"]:
            max_width["C"] = len(value["Ping"])
        excel_sheet["D" + str(excel_line)] = hostport
        excel_sheet["D" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(str(hostport)) > max_width["D"]:
            max_width["D"] = len(str(hostport))
        excel_sheet["E" + str(excel_line)] = value["Listening"]
        excel_sheet["E" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["Listening"]) > max_width["E"]:
            max_width["E"] = len(value["Listening"])
        excel_sheet["F" + str(excel_line)] = value["Message"]
        excel_sheet["F" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["Message"]) > max_width["F"]:
            max_width["F"] = len(value["Message"])
        excel_sheet["G" + str(excel_line)] = value["Listen:80"]
        excel_sheet["G" + str(excel_line)].font = openpyxl.styles.Font(color=color)
        if len(value["Listen:80"]) > max_width["G"]:
            max_width["G"] = len(value["Listen:80"])

    excel_sheet.column_dimensions["A"].width = max_width["A"] + 1
    excel_sheet.column_dimensions["B"].width = max_width["B"] + 1
    excel_sheet.column_dimensions["C"].width = max_width["C"] + 1
    excel_sheet.column_dimensions["D"].width = max_width["D"] + 1
    excel_sheet.column_dimensions["E"].width = max_width["E"] + 1
    excel_sheet.column_dimensions["F"].width = max_width["F"] + 1
    excel_sheet.column_dimensions["G"].width = max_width["G"] + 1

    excel_workbook.save(excel_filename)


####################################################################################################
def get_new_names(certs, errors):
    """ Return lists of common and alt names unmentioned in the input file """
    hostnames = []
    common_names = []
    alt_names = []

    # build lists of unique values
    for key, value in certs.items():
        hostname = re.sub(r":.*$", "", key)
        if hostname not in hostnames:
            hostnames.append(hostname)
        if value is not None:
            if 'subject' in value:
                if 'commonName' in value['subject']:
                    if value['subject']['commonName'] not in common_names:
                        common_names.append(value['subject']['commonName'])
            if 'extensions' in value:
                if 'subjectAltName' in value['extensions']:
                    for alt_name in value['extensions']['subjectAltName']:
                        if alt_name not in alt_names:
                            alt_names.append(alt_name)
    for key in errors.keys():
        hostname = re.sub(r":.*$", "", key)
        if hostname not in hostnames:
            hostnames.append(hostname)

    # build lists of new values
    new_common_names = []
    for common_name in common_names:
        if common_name not in hostnames:
            new_common_names.append(common_name)
    new_alt_names = []
    for alt_name in alt_names:
        if alt_name not in hostnames and alt_name not in common_names:
            new_alt_names.append(alt_name)

    return new_common_names, new_alt_names


####################################################################################################
def print_new_names(new_common_names, new_alt_names):
    """ Print lists of common and alt names unmentioned in the input file """
    if new_common_names:
        print("\nCommon names unmentioned in your input files:")
        cn = prettytable.PrettyTable()
        cn.set_style(prettytable.SINGLE_BORDER)
        cn.add_column("common name", new_common_names)
        cn.align = "l"
        print(cn)

    if new_alt_names:
        print("\nAlt names unmentioned in your input files:")
        an = prettytable.PrettyTable()
        an.set_style(prettytable.SINGLE_BORDER)
        an.add_column("alt name", new_alt_names)
        an.align = "l"
        print(an)


####################################################################################################
def update_excel(excel_filename, new_common_names, new_alt_names):
    """ Update an Excel workbook with tabs for common and alt names unmentioned in the input file """
    excel_workbook = openpyxl.load_workbook(excel_filename)

    # New common names
    excel_workbook.create_sheet(index=2, title="New common names")
    excel_sheet = excel_workbook.get_sheet_by_name("New common names")
    excel_sheet.page_setup.orientation = excel_sheet.ORIENTATION_PORTRAIT
    excel_line = 1
    max_width = {}

    excel_sheet["A1"] = "New common names"
    excel_sheet["A1"].font = openpyxl.styles.Font(bold=True)
    max_width["A"] = len("New common names")
    excel_sheet.freeze_panes = "A2"

    for name in new_common_names:
        excel_line += 1
        excel_sheet["A" + str(excel_line)] = name
        if len(name) > max_width["A"]:
            max_width["A"] = len(name)

    excel_sheet.column_dimensions["A"].width = max_width["A"] + 1

    # New alt names
    excel_workbook.create_sheet(index=3, title="New alt names")
    excel_sheet = excel_workbook.get_sheet_by_name("New alt names")
    excel_sheet.page_setup.orientation = excel_sheet.ORIENTATION_PORTRAIT
    excel_line = 1
    max_width = {}

    excel_sheet["A1"] = "New alt names"
    excel_sheet["A1"].font = openpyxl.styles.Font(bold=True)
    max_width["A"] = len("New alt names")
    excel_sheet.freeze_panes = "A2"

    for name in new_alt_names:
        excel_line += 1
        excel_sheet["A" + str(excel_line)] = name
        if len(name) > max_width["A"]:
            max_width["A"] = len(name)

    excel_sheet.column_dimensions["A"].width = max_width["A"] + 1

    # Save updated workbook
    excel_workbook.save(excel_filename)
